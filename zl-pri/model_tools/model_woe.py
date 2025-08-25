"""
Created on Sun Feb 24 19:44:03 2019

@author: zhoukun

email: 599518814@qq.com

"""
#%%
import pandas as pd
import numpy as np
import math
from openpyxl import load_workbook
import warnings
warnings.filterwarnings("ignore")

class Node(object):

    def __init__(self,vartype='num',left=None,right=None,
                 group_num=0,group_pct=0,good_num=0,good_pct=0,bad_num=0,bad_pct=0,
                 bad_rate=0,woe=0,iv=0):
        woep=float("{:.6f}".format(woe))
        if vartype=='num1':
            self.bin="({}, {}]".format(left,right)
            self.bindict={'left':[left],'right':[right],'woe':[woep]}
           
        elif vartype=='num2':
            left1=[i for i in left if i not in right]
            right1=[j for j in right if j not in left]
            if len(left1)==1:
                self.bin="({}, {}]".format(left1[0],right1[0])
            else:
                self.bin="&".join(["({}, {}]".format(left1[i],right1[i]) for i in range(len(left1))])
            self.bindict={'left':left1,'right':right1,'woe':[woep]*len(left1)}
   
        else:    
            self.bin=left
            self.bindict={tp:woep for tp in left}

        self.group_num = group_num
        self.group_pct = group_pct
        self.good_num = good_num
        self.good_pct = good_pct
        self.bad_num = bad_num
        self.bad_pct = bad_pct
        self.bad_rate = bad_rate
        self.woe = woep
        self.iv = iv*100


#%%
def change_var_dtype(df,varlist,var_type=2):
    
    if var_type==1:
        for vname in varlist:     
            try:
                df[vname] = df[vname].astype('object')
            except Exception as e:
                print('[error]',vname,e)  
    else:  
        for vname in varlist:     
            try:
                df[vname] = df[vname].astype('float64')
            except Exception as e:
                print('[error]',vname,e)
        
    return df



def char_calulate_woe(df, var, left, bt, gt, mode=1):
    
    df = df[df[var].isin([left])] 
    bad_cnt = sum(df['target'])
    good_cnt = df.shape[0] - bad_cnt

    if bad_cnt == 0:
        return Node(vartype='char1', left=[left], right='step1',group_num=bad_cnt+good_cnt,
                    group_pct=(bad_cnt+good_cnt) / (bt+gt),
                    good_num=good_cnt,
                    good_pct=good_cnt / gt,
                    bad_num=0,
                    bad_pct=0,
                    bad_rate=0,woe=0,iv=0)

    if good_cnt == 0:
        return Node(vartype='char1',left=[left],right='step1',group_num=bad_cnt+good_cnt,
                    group_pct=(bad_cnt+good_cnt) / (bt+gt),
                    good_num=0,
                    good_pct=0,
                    bad_num=bad_cnt,
                    bad_pct=bad_cnt / bt,
                    bad_rate=1,woe=0,iv=0)


    bad_rate=bad_cnt / (bad_cnt+good_cnt)
    br = bad_cnt / bt
    gr = good_cnt / gt
    woe = np.log(br / gr)
    iv = (br - gr) * woe

    return Node(vartype='char1',left=[left],right='step1',group_num=bad_cnt+good_cnt,
                group_pct=(bad_cnt+good_cnt) / (bt+gt),
                good_num=good_cnt,
                good_pct=good_cnt / gt,
                bad_num=bad_cnt,
                bad_pct=bad_cnt / bt,
                bad_rate=bad_rate,woe=woe,iv=iv)



def num_calulate_woe(df, var, left, right, bt, gt):
    
    df = df[(df[var]<=right)&(df[var]>left)]

    bad_cnt = sum(df['target'])
    good_cnt = df.shape[0] - bad_cnt

    if bad_cnt == 0:
        return Node(vartype='num1', left=left,right=right,group_num=bad_cnt+good_cnt,
                    group_pct=(bad_cnt+good_cnt) / (bt+gt),
                    good_num=good_cnt,
                    good_pct=good_cnt / gt,
                    bad_num=0,
                    bad_pct=0,
                    bad_rate=0,woe=0,iv=0)

    if good_cnt == 0:
        return Node(vartype='num1',left=left,right=right,group_num=bad_cnt+good_cnt,
                    group_pct=(bad_cnt+good_cnt) / (bt+gt),
                    good_num=0,
                    good_pct=0,
                    bad_num=bad_cnt,
                    bad_pct=bad_cnt / bt,
                    bad_rate=1,woe=0,iv=0)


    bad_rate=bad_cnt / (bad_cnt+good_cnt)
    br = bad_cnt / bt
    gr = good_cnt / gt
    woe = np.log(br / gr)
    iv = (br - gr) * woe

    return Node(vartype='num1', left=left,right=right,group_num=bad_cnt+good_cnt,
                group_pct=(bad_cnt+good_cnt) / (bt+gt),
                good_num=good_cnt,
                good_pct=good_cnt / gt,
                bad_num=bad_cnt,
                bad_pct=bad_cnt / bt,
                bad_rate=bad_rate,woe=woe,iv=iv)



def num_var_format(df, num, label, splitlist, bt, gt):
    
    infolist=[]
    for i in range(len(splitlist)):
        if i==0:
            nod=num_calulate_woe(df, var=num, left=float('-inf'), right=splitlist[i], bt=bt, gt=gt)
        else:
            nod=num_calulate_woe(df, var=num, left=splitlist[i-1], right=splitlist[i], bt=bt, gt=gt)
        infodict={
            'No':i+1,
            'varname':num,
            'label':label,
            'bin':nod.bin,
            'bindict':nod.bindict,
            'group_num':nod.group_num,
            'group_pct':nod.group_pct,
            'good_num':nod.good_num,
            'good_pct':nod.good_pct,
            'bad_num':nod.bad_num,
            'bad_pct':nod.bad_pct,
            'bad_rate':nod.bad_rate,
            'woe':nod.woe,
            'iv':nod.iv}
        infolist.append(infodict)

    infodf=pd.DataFrame(infolist,columns=['No','varname','label','bin','bindict',
                                          'good_num','good_pct','bad_num','bad_pct',
                                          'group_num','group_pct','bad_rate','woe','iv'])
    infodf['tot_iv']=infodf['iv'].sum()
    
    return infodf



def char_var_format(df, char, label, splitlist, bt, gt):
    
    infolist=[]
    for i,split in enumerate(splitlist):
        nod=char_calulate_woe(df, var=char, left=split, bt=bt, gt=gt)
        infodict={
                'No':i+1,
                'varname':char,
                'label':label,
                'bin':nod.bin,
                'bindict':nod.bindict,
                'group_num':nod.group_num,
                'group_pct':nod.group_pct,
                'good_num':nod.good_num,
                'good_pct':nod.good_pct,
                'bad_num':nod.bad_num,
                'bad_pct':nod.bad_pct,
                'bad_rate':nod.bad_rate,
                'woe':nod.woe,
                'iv':nod.iv}
        infolist.append(infodict)

    infodf=pd.DataFrame(infolist,columns=['No','varname','label','bin','bindict',
                                          'good_num','good_pct','bad_num','bad_pct',
                                          'group_num','group_pct','bad_rate','woe','iv'])
    infodf['tot_iv']=infodf['iv'].sum()

    return infodf



def woe_step1_outxlsx(data, num, char, label, file, bt, gt, listsp, n):

    writer=pd.ExcelWriter(file,engine ='xlsxwriter')
    workbook=writer.book
    format1=workbook.add_format({'align':'left'})
    format2=workbook.add_format({'align':'right','num_format':'0.00'})
    format3=workbook.add_format({'align':'right','num_format':'#,##0'})
    format4=workbook.add_format({'align':'right','num_format':'0.00%'})
    format5 = workbook.add_format({'bg_color':'#FFC7CE'})
    format6=workbook.add_format({'align':'right','num_format':'0.000000'})
    
    r=r2=0
    for var in num:
        print(var)
        temp=data[var]
        listsp=list(set(temp)&set(listsp))
        temp=temp[~temp.isin(listsp)]
        if len(temp)>0:
            splist=list(temp.quantile([j/n for j in range(n)]))
            splist=list(set([float("{:.5f}".format(i)) if len(str(math.modf(i)[0]))>7 else i for i in splist]+[temp.max()]))
        else:
            splist=[]
        splist=listsp+splist
        splist.sort()
        infodf=num_var_format(df=data[['target']+[var]],num=var, label=label[var],splitlist=splist,bt=bt,gt=gt)
        infoiv=infodf[['varname','tot_iv','label']].drop_duplicates()
        if r==0:
            infoiv.to_excel(writer,startrow=r,index=False,sheet_name='IV_TOTAL',encoding='utf-8')
        else:
            infoiv.to_excel(writer,startrow=r+1,header=0,index=False,sheet_name='IV_TOTAL',encoding='utf-8')
        infodf.to_excel(writer,startrow=r2,index=False,sheet_name='WoE_NUM',encoding='utf-8')
        worksheet=writer.sheets['WoE_NUM']
        worksheet.conditional_format(r2+1,12,r2+len(infodf)+1,12,{'type':'data_bar','data_bar_2010':True})
        worksheet.conditional_format(r2+1,14,r2+len(infodf)+1,14,{'type':'cell','criteria':'>','value':1,'format':format5})
        r2=r2+len(infodf)+5
        r=r+1

    r1=0  
    for var in char:
        print(var)
        splist=list(set(data[var]))
        infodf=char_var_format(df=data[['target']+[var]],char=var,label=label[var],splitlist=splist,bt=bt,gt=gt)
        infoiv=infodf[['varname','tot_iv','label']].drop_duplicates()
        if r==0:
            infoiv.to_excel(writer,startrow=r,index=False,sheet_name='IV_TOTAL',encoding='utf-8')
        else:
            infoiv.to_excel(writer,startrow=r+1,header=0,index=False,sheet_name='IV_TOTAL',encoding='utf-8')
        infodf.to_excel(writer,startrow=r1,index=False,sheet_name='WoE_CHAR',encoding='utf-8')  
        worksheet=writer.sheets['WoE_CHAR']
        worksheet.conditional_format(r1+1,12,r1+len(infodf)+1,12,{'type':'data_bar','data_bar_2010':True})
        worksheet.conditional_format(r1+1,14,r1+len(infodf)+1,14,{'type':'cell','criteria':'>','value':1,'format':format5})
        r1=r1+len(infodf)+5
        r=r+1
    if len(num)+len(char)>0:
        worksheet=writer.sheets['IV_TOTAL']
        worksheet.hide_gridlines(option = 2)
        worksheet.set_column('B:B',12,format2)
        worksheet.conditional_format(1,1,r,1,{'type':'data_bar','bar_solid':True,'bar_color':'#FFC7CE'})
        worksheet.set_column('A:A',12,format1)
        worksheet.set_column('C:C',12,format1)
    if len(num)>0:
        worksheet=writer.sheets['WoE_NUM']
        worksheet.hide_gridlines(option = 2)
        worksheet.set_column('A:A',3,format1)
        worksheet.set_column('L:L',10,format4)
        worksheet.set_column('M:M',10,format6)
        worksheet.set_column('N:O',10,format2)
        worksheet.set_column('F:F',10,format3)
        worksheet.set_column('H:H',10,format3)
        worksheet.set_column('J:J',10,format3)
        worksheet.set_column('G:G',10,format4)
        worksheet.set_column('I:I',10,format4)
        worksheet.set_column('K:K',10,format4)
        worksheet.set_column('B:D',10,format1)
        worksheet.set_column('E:E',10,format1,{'hidden':1})
    if len(char)>0:
        worksheet=writer.sheets['WoE_CHAR']
        worksheet.hide_gridlines(option = 2)
        worksheet.set_column('A:A',3,format1)
        worksheet.set_column('L:L',10,format4)
        worksheet.set_column('M:M',10,format6)
        worksheet.set_column('N:O',10,format2)
        worksheet.set_column('F:F',10,format3)
        worksheet.set_column('H:H',10,format3)
        worksheet.set_column('J:J',10,format3)
        worksheet.set_column('G:G',10,format4)
        worksheet.set_column('I:I',10,format4)
        worksheet.set_column('K:K',10,format4)
        worksheet.set_column('B:D',10,format1)
        worksheet.set_column('E:E',10,format1,{'hidden':1})
    writer.save()



#%%
###############################################################################

##################################new scheme###################################

###############################################################################



def woe_calulate(vartype, left, right, bad_cnt, good_cnt, bt, gt):

    if bad_cnt == 0:
        return Node(vartype=vartype, left=left, right=right,          
                    group_num=bad_cnt+good_cnt,
                    group_pct=(bad_cnt+good_cnt) / (bt+gt),
                    good_num=good_cnt,
                    good_pct=good_cnt / gt,
                    bad_num=0,
                    bad_pct=0,
                    bad_rate=float('inf'), woe=0, iv=0)

    if good_cnt == 0:
        return Node(vartype=vartype, left=left, right=right,
                    group_num=bad_cnt+good_cnt,
                    group_pct=(bad_cnt+good_cnt) / (bt+gt),
                    good_num=0,
                    good_pct=0,
                    bad_num=bad_cnt,
                    bad_pct=bad_cnt / bt,
                    bad_rate=0, woe=0, iv=0)

    bad_rate=bad_cnt / (bad_cnt+good_cnt)
    br = bad_cnt / bt
    gr = good_cnt / gt
    woe = np.log(br / gr)
    iv = (br - gr) * woe

    return Node(vartype=vartype, left=left, right=right,
                group_num=bad_cnt+good_cnt,
                group_pct=(bad_cnt+good_cnt) / (bt+gt),
                good_num=good_cnt,
                good_pct=good_cnt / gt,
                bad_num=bad_cnt,
                bad_pct=bad_cnt / bt,
                bad_rate=bad_rate, woe=woe, iv=iv)



def woe_format(vardict, var, label, vartype, bt, gt):
    
    infolist=[]
    for i,vdict in enumerate(vardict):
        vdict=vardict[vdict]
        nod=woe_calulate(vartype=vartype, left=vdict['left'], right=vdict['right'], bad_cnt=vdict['bad_cnt'], good_cnt=vdict['good_cnt'], bt=bt, gt=gt)

        infodict={
            'No':i+1,
            'varname':var,
            'label':label,
            'bin':nod.bin,
            'bindict':nod.bindict,
            'group_num':nod.group_num,
            'group_pct':nod.group_pct,
            'good_num':nod.good_num,
            'good_pct':nod.good_pct,
            'bad_num':nod.bad_num,
            'bad_pct':nod.bad_pct,
            'bad_rate':nod.bad_rate,
            'woe':nod.woe,
            'iv':nod.iv}
        infolist.append(infodict)

    infodf=pd.DataFrame(infolist,columns=['No','varname','label','bin','bindict',
                                          'good_num','good_pct','bad_num','bad_pct',
                                          'group_num','group_pct','bad_rate','woe','iv'])
    infodf['tot_iv']=infodf['iv'].sum()

    return infodf

          

def woe_vardict(dfvar, var, mode):

    vardict={}
    xlist=dfvar['select'].drop_duplicates().tolist() 
    for clr in xlist:
        xbin={}
        temp=dfvar[dfvar['select']==clr]
        if mode=='num2':
            xlist=temp['bindict'].tolist()
            xbin['left'],xbin['right']=[],[]
            for xdict in xlist:
                xdict=eval(xdict.replace('inf',"float('inf')"))
                xbin['left']=xbin['left']+xdict['left']
                xbin['right']=xbin['right']+xdict['right']
        else:
            xbin['left']=[j for i in temp['bin'].tolist() for j in eval(i)]
            xbin['right']='step2'
        xbin['bad_cnt']=sum(temp['bad_num'])
        xbin['good_cnt']=sum(temp['good_num'])
        vardict[clr]=xbin

    return vardict



def woe_outxlsx(data,num,char,label,file,bt,gt,step=1):

    writer=pd.ExcelWriter(file,engine ='xlsxwriter')
    workbook=writer.book
    format1=workbook.add_format({'align':'left'})
    format2=workbook.add_format({'align':'right','num_format':'0.00'})
    format3=workbook.add_format({'align':'right','num_format':'#,##0'})
    format4=workbook.add_format({'align':'right','num_format':'0.00%'})
    format5 = workbook.add_format({'bg_color':'#FFC7CE'})
    format6=workbook.add_format({'align':'right','num_format':'0.000000'})
    format7=workbook.add_format({'align':'center','num_format':'0'})
    
    r=r1=r2=0
    for var in num:
        if step==1:
            vardict=woe_vardict(dfvar=data[['target']+[var]], var=var, mode='num1')
            infodf=woe_format(vardict=vardict, var=var, label=label[var], vartype='num1', bt=bt, gt=gt)
        elif step==2:
            vardict=woe_vardict(dfvar=data[data['varname']==var], var=var, mode='num2')
            infodf=woe_format(vardict=vardict, var=var, label=label[var], vartype='num2', bt=bt, gt=gt) 
        else:
            infodf=num[var]

        infoiv=infodf[['varname','tot_iv','label']].drop_duplicates()
        if r==0:
            infoiv.to_excel(writer,startrow=r,index=False,sheet_name='IV_TOTAL',encoding='utf-8')
        else:
            infoiv.to_excel(writer,startrow=r+1,header=0,index=False,sheet_name='IV_TOTAL',encoding='utf-8')
        infodf.to_excel(writer,startrow=r1,index=False,sheet_name='WoE_NUM',encoding='utf-8')
        worksheet=writer.sheets['WoE_NUM']
        worksheet.conditional_format(r1+1,12,r1+len(infodf)+1,12,{'type':'data_bar','data_bar_2010':True})
        worksheet.conditional_format(r1+1,14,r1+len(infodf)+1,14,{'type':'cell','criteria':'>','value':1,'format':format5})
        r1=r1+len(infodf)+5
        r=r+1
        
    for var in char:
        if step==1:
            vardict=woe_vardict(dfvar=data[['target']+[var]], var=var, mode='char1')
            infodf=woe_format(vardict=vardict, var=var, label=label[var], vartype='char', bt=bt, gt=gt)
        elif step==2:
            vardict=woe_vardict(dfvar=data[data['varname']==var], var=var, mode='char2')
            infodf=woe_format(vardict=vardict, var=var, label=label[var], vartype='char', bt=bt, gt=gt)
        else:
            infodf=char[var]
        
        infoiv=infodf[['varname','tot_iv','label']].drop_duplicates()
        if r==0:
            infoiv.to_excel(writer,startrow=r,index=False,sheet_name='IV_TOTAL',encoding='utf-8')
        else:
            infoiv.to_excel(writer,startrow=r+1,header=0,index=False,sheet_name='IV_TOTAL',encoding='utf-8')
        infodf.to_excel(writer,startrow=r2,index=False,sheet_name='WoE_CHAR',encoding='utf-8')  
        worksheet=writer.sheets['WoE_CHAR']
        worksheet.conditional_format(r2+1,12,r2+len(infodf)+1,12,{'type':'data_bar','data_bar_2010':True})
        worksheet.conditional_format(r2+1,14,r2+len(infodf)+1,14,{'type':'cell','criteria':'>','value':1,'format':format5})
        r2=r2+len(infodf)+5
        r=r+1
        
    if len(num)+len(char)>0:
        worksheet=writer.sheets['IV_TOTAL']
        worksheet.hide_gridlines(option = 2)
        worksheet.set_column('B:B',12,format2)
        worksheet.conditional_format(1,1,r,1,{'type':'data_bar','bar_solid':True,'bar_color':'#FFC7CE'})
        worksheet.set_column('A:A',12,format1)
        worksheet.set_column('C:C',12,format1)
    if len(num)>0:
        worksheet=writer.sheets['WoE_NUM']
        worksheet.hide_gridlines(option = 2)
        worksheet.set_column('A:A',3,format1)
        worksheet.set_column('L:L',10,format4)
        worksheet.set_column('M:M',10,format6)
        worksheet.set_column('N:O',10,format2)
        worksheet.set_column('P:P',10,format7)
        worksheet.set_column('F:F',10,format3)
        worksheet.set_column('H:H',10,format3)
        worksheet.set_column('J:J',10,format3)
        worksheet.set_column('G:G',10,format4)
        worksheet.set_column('I:I',10,format4)
        worksheet.set_column('K:K',10,format4)
        worksheet.set_column('B:D',10,format1)
        worksheet.set_column('E:E',10,format1,{'hidden':1})
    if len(char)>0:
        worksheet=writer.sheets['WoE_CHAR']
        worksheet.hide_gridlines(option = 2)
        worksheet.set_column('A:A',3,format1)
        worksheet.set_column('L:L',10,format4)
        worksheet.set_column('M:M',10,format6)
        worksheet.set_column('N:O',10,format2)
        worksheet.set_column('P:P',10,format7)
        worksheet.set_column('F:F',10,format3)
        worksheet.set_column('H:H',10,format3)
        worksheet.set_column('J:J',10,format3)
        worksheet.set_column('G:G',10,format4)
        worksheet.set_column('I:I',10,format4)
        worksheet.set_column('K:K',10,format4)
        worksheet.set_column('B:D',10,format1)
        worksheet.set_column('E:E',10,format1,{'hidden':1})        
    writer.save()



def woe_readxlsx(file, stname, keep=False):

    temp=pd.read_excel(file,sheet_name=stname)
    temp=temp[(temp['varname']!='varname')&(temp['varname'].notnull())]
    temp=temp.reset_index()
    temp['index']=temp['index'].map(lambda x: x+2)
    rowlist=temp['index'].tolist()
    wb=load_workbook(file)
    ws=wb[stname]
    temp=pd.DataFrame()
    for row in rowlist:
        rowinfo={}
        c1=ws.cell(row,2)
        rowinfo['varname']=[c1.value]
        c2=ws.cell(row,3)
        rowinfo['label']=[c2.value]
        c3=ws.cell(row,4)
        rowinfo['bin']=[c3.value]        
        c4=ws.cell(row,5)
        rowinfo['bindict']=[c4.value]
        c5=ws.cell(row,6)
        rowinfo['good_num']=[c5.value]
        c6=ws.cell(row,8)
        rowinfo['bad_num']=[c6.value]
        c7=ws.cell(row,13)
        rowinfo['select']=[str(c7.fill.start_color.rgb)[:8]+'_'+str(c7.fill.start_color.theme)[:1]+"{:.3f}".format(c7.fill.start_color.tint)]
        rowinfo=pd.DataFrame(rowinfo)
        temp=temp.append(rowinfo)
    listvar=temp['varname'].drop_duplicates().tolist() 
    if keep==False:
        for var in listvar:
            temp1=temp[temp['varname']==var]
            if len(temp1['select'].drop_duplicates())==1:
                temp=temp[temp['varname']!=var]
    temp=temp.reset_index(drop=True)
    listvar=temp['varname'].drop_duplicates().tolist()    
    for var in listvar:
        temp1=temp[temp['varname']==var]
        j,chk,chd=1,[],{}
        for i in temp1.index.tolist():
            if temp1.loc[i,'select']=='00000000_V0.000':
                temp.loc[i,'select']='color'+str(j)
                j=j+1
            else:
                if temp1.loc[i,'select'] in chk:
                    temp.loc[i,'select'] =chd[temp1.loc[i,'select']]
                else:
                    temp.loc[i,'select'] ='color'+str(j)
                    chk=chk+[temp1.loc[i,'select']]
                    chd[temp1.loc[i,'select']]='color'+str(j)
                    j=j+1                  
    return temp



def woe_step1(data, target='target', varconf=None, fileout='WoE_Step1.xlsx', compare_check=False, listsp=[-9999999,-9999998], n=20):
    
    if target!='target':
        data=data.copy(deep=True).rename(columns={target:'target'})
    if len(data['target'])>data['target'].count():
        print("e1:目标变量存在缺失值！")
        return        
    if len(set(data['target']))>2:
        print("e2:目标变量不是二分类变量！")
        return
    if sum(data['target'])>len(data)*0.5:
        print("e3:请确认目标变量{good:0,bad:1}定义无误！")
    
    bt=sum(data['target'])
    gt=len(data)-bt
    
    if varconf is None:
        num=data.drop('target',axis=1).select_dtypes('number').columns.tolist()
        char=data.drop('target',axis=1).select_dtypes(['object','category']).columns.tolist()
        label={var:'' for var in num+char}
    elif type(varconf) is dict:
        num=data.drop('target',axis=1).select_dtypes('number').columns.tolist()
        char=data.drop('target',axis=1).select_dtypes(['object','category']).columns.tolist()
        label=varconf        
    else:
        varconf=varconf[varconf['type']!=0]         
        label=varconf[['colname','label']].set_index('colname').to_dict()['label']       
        num=varconf[varconf['type']==2]['colname'].tolist()
        numchar=data[num].select_dtypes('number').columns
        numchar=[var for var in num if var not in numchar]
        data=change_var_dtype(data,varlist=numchar,var_type=2)
        char=varconf[varconf['type']==1]['colname'].tolist()
        data=change_var_dtype(data,varlist=char,var_type=1)
    if compare_check==False:
        woe_step1_outxlsx(data=data, num=num, char=char, label=label, file=fileout, bt=bt, gt=gt, listsp=listsp, n=n)
    else:
        return data,num,char,label,gt,bt


      
def woe_step2(filein, fileout, varconf=None, keep=False):

    try:
        numdf=pd.read_excel(filein,sheet_name='WoE_NUM')
    except Exception as e:
        print('numvar read [error]',e)
        numdf=pd.DataFrame()
    try:
        chardf=pd.read_excel(filein,sheet_name='WoE_CHAR')
    except Exception as e:
        print('chardf read [error]',e)
        chardf=pd.DataFrame()
    if len(numdf)>0:
        numdf=woe_readxlsx(file=filein, stname='WoE_NUM', keep=keep)
        num=numdf['varname'].drop_duplicates().tolist()
    else:
        num=[]
    if len(chardf)>0:
        chardf=woe_readxlsx(file=filein,stname='WoE_CHAR', keep=keep)
        char=chardf['varname'].drop_duplicates().tolist()
    else:
        char=[]
    alldf=numdf.append(chardf)
    
    if varconf is None:
        label=pd.read_excel(filein,sheet_name='IV_TOTAL')
        label=label[['varname','label']].set_index('varname')
        label=label.to_dict()['label']
    else:       
        label=varconf[['varname','label']].set_index('varname')
        label=label.to_dict()['label']
    
    cal=alldf[['varname','good_num','bad_num']]
    cal=cal.groupby('varname')['good_num','bad_num'].sum()
    cal=cal.drop_duplicates()
    if len(cal)>=2:
        print("e4:变量 %s 好坏样本不一致，请检查数据！"%(str(cal.index.tolist())))
        return
    else:
        gt=cal['good_num'].max()
        bt=cal['bad_num'].max()

    woe_outxlsx(data=alldf, num=num, char=char, label=label, file=fileout, bt=bt, gt=gt ,step=2)



def woe_compare(train, test, oot, target='target', varconf=None, fileout='WoE_Compare.xlsx', listsp=[-9999999,-9999998], n=20):
    
    train,num,char,label,gt1,bt1=woe_step1(data=train, target=target, varconf=varconf, compare_check=True)
    test,_,_,_,gt2,bt2=woe_step1(data=test, target=target, varconf=varconf, compare_check=True)
    oot,_,_,_,gt3,bt3=woe_step1(data=oot, target=target, varconf=varconf, compare_check=True)
    if len(train)==0 or len(test)==0 or len(oot)==0:
        return

    writer=pd.ExcelWriter(fileout,engine ='xlsxwriter')
    workbook=writer.book
    format1=workbook.add_format({'align':'left'})
    format2=workbook.add_format({'align':'right','num_format':'0.0'})
    format3=workbook.add_format({'align':'right','num_format':'#,##0'})
    format4=workbook.add_format({'align':'right','num_format':'0.0%'})
    format5 = workbook.add_format({'bg_color':'#FABF8F'})
    format6=workbook.add_format({'align':'right','num_format':'0.000000'})
    
    r,r1,r2=0,0,0
    for var in num:
        print(var)
        temp=train[var]
        listch=list(set(temp)&set(listsp))
        temp=temp[~temp.isin(listsp)]
        if len(temp)>0:
            splist=list(temp.quantile([j/n for j in range(n)]))
            splist=list(set([float("{:.5f}".format(i)) if len(str(math.modf(i)[0]))>7 else i for i in splist]+[temp.max()]))
        else:
            splist=[]
        splist=listch+splist
        splist.sort()
        infodf1=num_var_format(df=train[['target']+[var]],num=var, label=label[var],splitlist=splist,bt=bt1,gt=gt1)
        infoiv=infodf1[['varname','tot_iv','label']].drop_duplicates()
        splist.pop(-1)
        splist=splist+[test[var].max()]
        infodf2=num_var_format(df=test[['target']+[var]],num=var, label=label[var],splitlist=splist,bt=bt2,gt=gt2)
        splist.pop(-1)
        splist=splist+[oot[var].max()]        
        infodf3=num_var_format(df=oot[['target']+[var]],num=var, label=label[var],splitlist=splist,bt=bt3,gt=gt3)         
        if r==0:
            infoiv.to_excel(writer,startrow=r,index=False,sheet_name='IV_TOTAL',encoding='utf-8')
        else:
            infoiv.to_excel(writer,startrow=r+1,header=0,index=False,sheet_name='IV_TOTAL',encoding='utf-8')
        infodf1.drop(['iv','tot_iv'],axis=1).to_excel(writer,startrow=r2,startcol=0,index=False,sheet_name='WoE_NUM',encoding='utf-8')
        infodf2[['woe']].rename(columns={'woe':'woe_test'}).to_excel(writer,startrow=r2,startcol=13,index=False,sheet_name='WoE_NUM',encoding='utf-8')
        infodf3[['woe']].rename(columns={'woe':'woe_oot'}).to_excel(writer,startrow=r2,startcol=14,index=False,sheet_name='WoE_NUM',encoding='utf-8')
        infodf1[['iv','tot_iv']].to_excel(writer,startrow=r2,startcol=15,index=False,sheet_name='WoE_NUM',encoding='utf-8')   
        infodf2.to_excel(writer,startrow=r2,startcol=19,index=False,sheet_name='WoE_NUM',encoding='utf-8')
        infodf3.to_excel(writer,startrow=r2,startcol=36,index=False,sheet_name='WoE_NUM',encoding='utf-8')
        worksheet=writer.sheets['WoE_NUM']
        worksheet.add_table(r2+1, 0, r2+len(infodf1), 16, {'header_row': 0, 'style':'Table Style Light 0'})
        worksheet.add_table(r2+1, 19, r2+len(infodf1), 33, {'header_row': 0, 'style':'Table Style Light 0'})
        worksheet.add_table(r2+1, 36, r2+len(infodf1), 50, {'header_row': 0, 'style':'Table Style Light 0'})
        worksheet.conditional_format(r2+1,10,r2+len(infodf1),10,{'type':'cell','criteria':'<','value':0.02,'format':format5})
        worksheet.conditional_format(r2+1,12,r2+len(infodf1),12,{'type':'data_bar','data_bar_2010':True})
        worksheet.conditional_format(r2+1,13,r2+len(infodf1),13,{'type':'data_bar','data_bar_2010':True})
        worksheet.conditional_format(r2+1,14,r2+len(infodf1),14,{'type':'data_bar','data_bar_2010':True})
        worksheet.conditional_format(r2+1,16,r2+len(infodf1),16,{'type':'cell','criteria':'>','value':1,'format':format5})
        for i in range(2):
            worksheet.conditional_format(r2+1,29+17*i,r2+len(infodf1),29+17*i,{'type':'cell','criteria':'<','value':0.02,'format':format5})
            worksheet.conditional_format(r2+1,31+17*i,r2+len(infodf1),31+17*i,{'type':'data_bar','data_bar_2010':True})
            worksheet.conditional_format(r2+1,33+17*i,r2+len(infodf1),33+17*i,{'type':'cell','criteria':'>','value':1,'format':format5})
        r,r2=r+1,r2+len(infodf1)+5

    for var in char:
        splist=list(set(train[var]))
        infodf1=char_var_format(df=train[['target']+[var]],char=var,label=label[var],splitlist=splist,bt=bt1,gt=gt1)
        infoiv=infodf1[['varname','tot_iv','label']].drop_duplicates()
        infodf2=char_var_format(df=test[['target']+[var]],char=var,label=label[var],splitlist=splist,bt=bt2,gt=gt2)
        infodf3=char_var_format(df=oot[['target']+[var]],char=var,label=label[var],splitlist=splist,bt=bt3,gt=gt3)
        if r==0:
            infoiv.to_excel(writer,startrow=r,index=False,sheet_name='IV_TOTAL',encoding='utf-8')
        else:
            infoiv.to_excel(writer,startrow=r+1,header=0,index=False,sheet_name='IV_TOTAL',encoding='utf-8')
        infodf1.drop(['iv','tot_iv'],axis=1).to_excel(writer,startrow=r1,startcol=0,index=False,sheet_name='WoE_CHAR',encoding='utf-8')
        infodf2[['woe']].rename(columns={'woe':'woe_test'}).to_excel(writer,startrow=r1,startcol=13,index=False,sheet_name='WoE_CHAR',encoding='utf-8')
        infodf3[['woe']].rename(columns={'woe':'woe_oot'}).to_excel(writer,startrow=r1,startcol=14,index=False,sheet_name='WoE_CHAR',encoding='utf-8')
        infodf1[['iv','tot_iv']].to_excel(writer,startrow=r1,startcol=15,index=False,sheet_name='WoE_CHAR',encoding='utf-8')
        infodf2.to_excel(writer,startrow=r1,startcol=19,index=False,sheet_name='WoE_CHAR',encoding='utf-8')
        infodf3.to_excel(writer,startrow=r1,startcol=36,index=False,sheet_name='WoE_CHAR',encoding='utf-8')        
        worksheet=writer.sheets['WoE_CHAR']
        worksheet.add_table(r1+1, 0, r1+len(infodf1), 16, {'header_row': 0, 'style':'Table Style Light 0'})
        worksheet.add_table(r1+1, 19, r1+len(infodf1), 33, {'header_row': 0, 'style':'Table Style Light 0'})
        worksheet.add_table(r1+1, 36, r1+len(infodf1), 50, {'header_row': 0, 'style':'Table Style Light 0'})
        worksheet.conditional_format(r1+1,10,r1+len(infodf1),10,{'type':'cell','criteria':'<','value':0.02,'format':format5})
        worksheet.conditional_format(r1+1,12,r1+len(infodf1),12,{'type':'data_bar','data_bar_2010':True})
        worksheet.conditional_format(r1+1,13,r1+len(infodf1),13,{'type':'data_bar','data_bar_2010':True})
        worksheet.conditional_format(r1+1,14,r1+len(infodf1),14,{'type':'data_bar','data_bar_2010':True})
        worksheet.conditional_format(r1+1,16,r1+len(infodf1),16,{'type':'cell','criteria':'>','value':1,'format':format5})
        for i in range(2):
            worksheet.conditional_format(r1+1,29+17*i,r1+len(infodf1),29+17*i,{'type':'cell','criteria':'<','value':0.02,'format':format5})
            worksheet.conditional_format(r1+1,31+17*i,r1+len(infodf1),31+17*i,{'type':'data_bar','data_bar_2010':True})
            worksheet.conditional_format(r1+1,33+17*i,r1+len(infodf1),33+17*i,{'type':'cell','criteria':'>','value':1,'format':format5})
        r,r1=r+1,r1+len(infodf1)+5

    if len(num)+len(char)>0:
        worksheet=writer.sheets['IV_TOTAL']
        worksheet.freeze_panes(1,0) 
        worksheet.hide_gridlines(option = 2)
        worksheet.conditional_format(1,1,r,1,{'type':'data_bar','bar_solid':True,'bar_color':'#FABF8F'})
        worksheet.set_column(0,0,12,format1)
        worksheet.set_column(1,1,12,format2)
        worksheet.set_column(2,2,12,format1)
    if len(num)>0:
        worksheet=writer.sheets['WoE_NUM']
        worksheet.hide_gridlines(option = 2)
        worksheet.set_column(0,0,3,format1)
        worksheet.set_column(1,3,10,format1)
        worksheet.set_column(4,4,10,format1,{'hidden':1})
        worksheet.set_column(5,5,10,format3)
        worksheet.set_column(6,6,10,format4)
        worksheet.set_column(7,7,10,format3)
        worksheet.set_column(8,8,10,format4)
        worksheet.set_column(9,9,10,format3)
        worksheet.set_column(10,10,10,format4)
        worksheet.set_column(11,11,10,format4)
        worksheet.set_column(12,14,10,format6)
        worksheet.set_column(15,16,10,format2)
        for i in range(2):
            worksheet.set_column(19+17*i,19+17*i,3,format1)
            worksheet.set_column(20+17*i,22+17*i,10,format1)
            worksheet.set_column(23+17*i,23+17*i,10,format1,{'hidden':1})
            worksheet.set_column(24+17*i,24+17*i,10,format3)
            worksheet.set_column(25+17*i,25+17*i,10,format4)
            worksheet.set_column(26+17*i,26+17*i,10,format3)
            worksheet.set_column(27+17*i,27+17*i,10,format4)
            worksheet.set_column(28+17*i,28+17*i,10,format3)
            worksheet.set_column(29+17*i,29+17*i,10,format4)
            worksheet.set_column(30+17*i,30+17*i,10,format4)
            worksheet.set_column(31+17*i,31+17*i,10,format6)
            worksheet.set_column(32+17*i,33+17*i,10,format2) 
    if len(char)>0:
        worksheet=writer.sheets['WoE_CHAR']
        worksheet.hide_gridlines(option = 2)
        worksheet.set_column(0,0,3,format1)
        worksheet.set_column(1,3,10,format1)
        worksheet.set_column(4,4,10,format1,{'hidden':1})
        worksheet.set_column(5,5,10,format3)
        worksheet.set_column(6,6,10,format4)
        worksheet.set_column(7,7,10,format3)
        worksheet.set_column(8,8,10,format4)
        worksheet.set_column(9,9,10,format3)
        worksheet.set_column(10,10,10,format4)
        worksheet.set_column(11,11,10,format4)
        worksheet.set_column(12,14,10,format6)
        worksheet.set_column(15,16,10,format2)
        for i in range(2):
            worksheet.set_column(19+17*i,19+17*i,3,format1)
            worksheet.set_column(20+17*i,22+17*i,10,format1)
            worksheet.set_column(23+17*i,23+17*i,10,format1,{'hidden':1})
            worksheet.set_column(24+17*i,24+17*i,10,format3)
            worksheet.set_column(25+17*i,25+17*i,10,format4)
            worksheet.set_column(26+17*i,26+17*i,10,format3)
            worksheet.set_column(27+17*i,27+17*i,10,format4)
            worksheet.set_column(28+17*i,28+17*i,10,format3)
            worksheet.set_column(29+17*i,29+17*i,10,format4)
            worksheet.set_column(30+17*i,30+17*i,10,format4)
            worksheet.set_column(31+17*i,31+17*i,10,format6)
            worksheet.set_column(32+17*i,33+17*i,10,format2) 

    writer.save()
#%%

def woe_quick(data, target='target', varconf=None, keepvar=['var_id'], fileout='WoE_Quick.xlsx'):

    data,num,char,label,gt,bt=woe_step1(data=data, target=target, varconf=varconf, compare_check=True)
    varconf["split"]=varconf["split"].str.replace("inf","float('inf')").map(lambda x: None if pd.isnull(x) else eval(x))
    split=varconf[varconf.colname.isin(num)].set_index("colname")["split"].to_dict()

    writer=pd.ExcelWriter(fileout,engine ='xlsxwriter')
    workbook=writer.book
    format1=workbook.add_format({'align':'left'})
    format2=workbook.add_format({'align':'right','num_format':'0.00'})
    format3=workbook.add_format({'align':'right','num_format':'#,##0'})
    format4=workbook.add_format({'align':'right','num_format':'0.00%'})
    format5 = workbook.add_format({'bg_color':'#FFC7CE'})
    format6=workbook.add_format({'align':'right','num_format':'0.000000'})
    format7=workbook.add_format({'align':'center','num_format':'0'})
    r=r1=r2=0
    for var in num:
        splitlist=split[var]
        infodf=num_var_format(df=data[['target']+[var]], num=var, label=label[var], splitlist=splitlist, bt=bt, gt=gt)
        infoiv=infodf[['varname','tot_iv','label']].drop_duplicates()
        if r==0:
            infoiv.to_excel(writer,startrow=r,index=False,sheet_name='IV_TOTAL',encoding='utf-8')
        else:
            infoiv.to_excel(writer,startrow=r+1,header=0,index=False,sheet_name='IV_TOTAL',encoding='utf-8')
        infodf.to_excel(writer,startrow=r1,index=False,sheet_name='WoE_NUM',encoding='utf-8')
        worksheet=writer.sheets['WoE_NUM']
        worksheet.conditional_format(r1+1,12,r1+len(infodf)+1,12,{'type':'data_bar','data_bar_2010':True})
        worksheet.conditional_format(r1+1,14,r1+len(infodf)+1,14,{'type':'cell','criteria':'>','value':1,'format':format5})
        r,r1=r+1,r1+len(infodf)+5

    for var in char:
        splitlist=list(set(data[var]))
        infodf=char_var_format(df=data[['target']+[var]],char=var,label=label[var],splitlist=splitlist,bt=bt,gt=gt)
        infoiv=infodf[['varname','tot_iv','label']].drop_duplicates()
        if r==0:
            infoiv.to_excel(writer,startrow=r,index=False,sheet_name='IV_TOTAL',encoding='utf-8')
        else:
            infoiv.to_excel(writer,startrow=r+1,header=0,index=False,sheet_name='IV_TOTAL',encoding='utf-8')
        infodf.to_excel(writer,startrow=r2,index=False,sheet_name='WoE_CHAR',encoding='utf-8')  
        worksheet=writer.sheets['WoE_CHAR']
        worksheet.conditional_format(r2+1,12,r2+len(infodf)+1,12,{'type':'data_bar','data_bar_2010':True})
        worksheet.conditional_format(r2+1,14,r2+len(infodf)+1,14,{'type':'cell','criteria':'>','value':1,'format':format5})
        r,r2=r+1,r2+len(infodf)+5
        
    if len(num)+len(char)>0:
        worksheet=writer.sheets['IV_TOTAL']
        worksheet.hide_gridlines(option = 2)
        worksheet.set_column('B:B',12,format2)
        worksheet.conditional_format(1,1,r,1,{'type':'data_bar','bar_solid':True,'bar_color':'#FFC7CE'})
        worksheet.set_column('A:A',12,format1)
        worksheet.set_column('C:C',12,format1)
    if len(num)>0:
        worksheet=writer.sheets['WoE_NUM']
        worksheet.hide_gridlines(option = 2)
        worksheet.set_column('A:A',3,format1)
        worksheet.set_column('L:L',10,format4)
        worksheet.set_column('M:M',10,format6)
        worksheet.set_column('N:O',10,format2)
        worksheet.set_column('P:P',10,format7)
        worksheet.set_column('F:F',10,format3)
        worksheet.set_column('H:H',10,format3)
        worksheet.set_column('J:J',10,format3)
        worksheet.set_column('G:G',10,format4)
        worksheet.set_column('I:I',10,format4)
        worksheet.set_column('K:K',10,format4)
        worksheet.set_column('B:D',10,format1)
        worksheet.set_column('E:E',10,format1,{'hidden':1})
    if len(char)>0:
        worksheet=writer.sheets['WoE_CHAR']
        worksheet.hide_gridlines(option = 2)
        worksheet.set_column('A:A',3,format1)
        worksheet.set_column('L:L',10,format4)
        worksheet.set_column('M:M',10,format6)
        worksheet.set_column('N:O',10,format2)
        worksheet.set_column('P:P',10,format7)
        worksheet.set_column('F:F',10,format3)
        worksheet.set_column('H:H',10,format3)
        worksheet.set_column('J:J',10,format3)
        worksheet.set_column('G:G',10,format4)
        worksheet.set_column('I:I',10,format4)
        worksheet.set_column('K:K',10,format4)
        worksheet.set_column('B:D',10,format1)
        worksheet.set_column('E:E',10,format1,{'hidden':1})        
    writer.save()

